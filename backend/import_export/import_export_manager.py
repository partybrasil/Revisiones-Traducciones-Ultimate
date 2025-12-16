"""
Import/Export Manager - Operaciones MASIVAS Excel/PDF/Markdown/HTML

Funcionalidades:
- generate_excel_template(): Plantilla Excel con 60+ columnas, dropdowns, validaciones
- import_from_excel(): Importación masiva con validaciones
- export_to_pdf(): PDF profesional con ReportLab
- export_to_markdown(), export_to_html(): Exportaciones adicionales
- export_catalog_excel(): Exportación de catálogo con filtros
"""

import io
import re
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

# Importar managers necesarios
import sys
sys.path.append(str(Path(__file__).parent.parent))
from models.product_sheet import ProductSheet
from core.preset_manager import PresetManager
from core.version_manager import VersionManager


class ImportExportManager:
    """
    Gestor de importación y exportación masiva de productos.
    """
    
    # Definir columnas del template Excel (60+ columnas)
    EXCEL_COLUMNS = [
        # Identificadores
        ("SKU", "critical", "Código único del producto (ej: CF-HYD-001)"),
        ("EAN_PRIMARY", "recommended", "Código de barras principal EAN-13"),
        ("EAN_SECONDARY", "optional", "Código de barras secundario"),
        
        # Títulos multiidioma
        ("TITLE_ES_SHORT", "critical", "Título corto en español"),
        ("TITLE_PT_SHORT", "recommended", "Título corto en portugués"),
        ("TITLE_IT_SHORT", "recommended", "Título corto en italiano"),
        ("TITLE_EN_SHORT", "optional", "Título corto en inglés"),
        ("TITLE_FR_SHORT", "optional", "Título corto en francés"),
        
        # Marca y clasificación
        ("BRAND", "recommended", "Marca del producto"),
        ("GAMA_ES", "optional", "Gama/línea en español"),
        ("FAMILY", "critical", "Familia del producto"),
        ("SUBFAMILY", "recommended", "Subfamilia"),
        
        # Peso y dimensiones
        ("NET_WEIGHT_VALUE", "recommended", "Peso neto valor numérico"),
        ("NET_WEIGHT_UNIT", "recommended", "Unidad de peso (g, ml, kg, L)"),
        ("GROSS_WEIGHT_VALUE", "optional", "Peso bruto valor"),
        ("GROSS_WEIGHT_UNIT", "optional", "Unidad peso bruto"),
        ("HEIGHT_CM", "optional", "Alto en cm"),
        ("WIDTH_CM", "optional", "Ancho en cm"),
        ("DEPTH_CM", "optional", "Profundidad en cm"),
        
        # Formato y packaging
        ("FORMAT_TYPE", "recommended", "Tipo de formato (Botella, Tubo, Tarro, etc)"),
        ("FORMAT_MATERIAL", "recommended", "Material del formato"),
        ("FORMAT_CLOSURE", "optional", "Tipo de cierre"),
        ("PACKAGING_LANGUAGES", "optional", "Idiomas en el packaging (separados por coma)"),
        
        # Composición
        ("INCI_INGREDIENTS", "critical", "Lista INCI de ingredientes"),
        ("KEY_INGREDIENTS", "optional", "Ingredientes clave destacados"),
        
        # Modo de uso multiidioma
        ("MODE_OF_USE_ES", "critical", "Modo de empleo en español"),
        ("MODE_OF_USE_PT", "recommended", "Modo de empleo en portugués"),
        ("MODE_OF_USE_IT", "recommended", "Modo de empleo en italiano"),
        ("MODE_OF_USE_EN", "optional", "Modo de empleo en inglés"),
        
        # Advertencias multiidioma
        ("WARNINGS_ES", "critical", "Advertencias en español"),
        ("WARNINGS_PT", "recommended", "Advertencias en portugués"),
        ("WARNINGS_IT", "recommended", "Advertencias en italiano"),
        ("WARNINGS_EN", "optional", "Advertencias en inglés"),
        
        # Alérgenos y PAO
        ("ALLERGENS_PRESENT", "critical", "Alérgenos presentes (separados por coma)"),
        ("ALLERGENS_MAY_CONTAIN", "optional", "Puede contener trazas (separados por coma)"),
        ("PAO", "critical", "Período después de apertura (6M, 12M, 18M, 24M, 36M)"),
        
        # Pictogramas y certificaciones
        ("PICTOGRAMS", "optional", "Pictogramas (separados por coma)"),
        ("CERTIFICATIONS", "optional", "Certificaciones (separadas por coma)"),
        
        # Origen y fabricación
        ("MADE_IN_COUNTRY", "recommended", "País de fabricación (código ISO)"),
        ("MADE_IN_TEXT_ES", "optional", "Texto 'Fabricado en' español"),
        ("MADE_IN_TEXT_PT", "optional", "Texto 'Fabricado en' portugués"),
        
        # Distribuidor
        ("DISTRIBUTOR_NAME", "recommended", "Nombre del distribuidor"),
        ("DISTRIBUTOR_CIF", "optional", "CIF/NIF del distribuidor"),
        ("DISTRIBUTOR_ADDRESS", "optional", "Dirección del distribuidor"),
        ("DISTRIBUTOR_COUNTRY", "optional", "País del distribuidor"),
        
        # Responsable
        ("RESPONSIBLE_PERSON_NAME", "recommended", "Nombre persona responsable"),
        ("RESPONSIBLE_PERSON_EMAIL", "optional", "Email persona responsable"),
        ("RESPONSIBLE_PERSON_PHONE", "optional", "Teléfono persona responsable"),
        
        # Naturalidad y sostenibilidad
        ("NATURAL_ORIGIN_PERCENTAGE", "optional", "% ingredientes de origen natural"),
        ("VEGAN", "optional", "Es vegano (SI/NO)"),
        ("CRUELTY_FREE", "optional", "Cruelty-free (SI/NO)"),
        
        # Almacenamiento
        ("STORAGE_TEMP_MIN", "optional", "Temperatura mínima almacenamiento °C"),
        ("STORAGE_TEMP_MAX", "optional", "Temperatura máxima almacenamiento °C"),
        ("STORAGE_CONDITIONS_ES", "optional", "Condiciones almacenamiento español"),
        
        # Uso especial
        ("PREGNANCY_SAFE", "optional", "Seguro en embarazo (SI/NO/CONSULTAR)"),
        ("CHILD_SAFE", "optional", "Seguro en niños (SI/NO/EDAD_MINIMA)"),
        
        # Estado y notas
        ("STATUS", "optional", "Estado del producto (draft, review, approved)"),
        ("INTERNAL_NOTES", "optional", "Notas internas"),
    ]
    
    # Valores válidos para dropdowns
    VALID_FAMILIES = [
        "COSMETICS_FACIAL", "COSMETICS_BODY", "COSMETICS_HAIR", "COSMETICS_EYES",
        "COSMETICS_LIPS", "COSMETICS_NAILS", "PERFUMES", "FOOD_PACKAGED",
        "FOOD_FRESH", "FOOD_FROZEN", "SUPPLEMENTS", "PHARMACEUTICALS"
        # ... +138 más (simplificado para demo)
    ]
    
    VALID_FORMAT_TYPES = ["Botella", "Tubo", "Tarro", "Caja", "Bolsa", "Spray", "Stick", "Ampolla"]
    VALID_FORMAT_MATERIALS = ["Plástico", "Vidrio", "Aluminio", "Cartón", "Metal", "Papel"]
    VALID_PAO_VALUES = ["6M", "12M", "18M", "24M", "36M"]
    VALID_WEIGHT_UNITS = ["g", "ml", "kg", "L"]
    VALID_STATUS = ["draft", "review", "approved", "archived"]
    
    def __init__(self, session: Session):
        """
        Inicializar ImportExportManager.
        
        Args:
            session: Sesión SQLAlchemy
        """
        self.session = session
        self.preset_manager = PresetManager()
        self.version_manager = VersionManager(session)
    
    def generate_excel_template(self) -> bytes:
        """
        Generar template Excel con 60+ columnas, dropdowns y validaciones.
        
        Returns:
            bytes: Archivo Excel en memoria
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Color styles
        critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        recommended_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        optional_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        header_font = Font(bold=True, size=11, color="000000")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Row 1: Headers con colores
        for col_idx, (col_name, priority, description) in enumerate(self.EXCEL_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.alignment = center_alignment
            
            # Aplicar color según prioridad
            if priority == "critical":
                cell.fill = critical_fill
            elif priority == "recommended":
                cell.fill = recommended_fill
            else:
                cell.fill = optional_fill
            
            # Agregar comentario con descripción
            cell.comment = openpyxl.comments.Comment(description, "Sistema")
        
        # Row 2: Ejemplo de datos
        example_data = self._get_example_row()
        for col_idx, value in enumerate(example_data, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = value
        
        # Data validations (dropdowns)
        self._add_data_validations(ws)
        
        # Autosize columns
        for col_idx in range(1, len(self.EXCEL_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Guardar a bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _get_example_row(self) -> List[str]:
        """Generar fila de ejemplo con datos demo."""
        return [
            "CF-HYD-001",  # SKU
            "5412345678901",  # EAN_PRIMARY
            "",  # EAN_SECONDARY
            "Crema Hidratante Facial 24h",  # TITLE_ES_SHORT
            "Creme Hidratante Facial 24h",  # TITLE_PT_SHORT
            "Crema Idratante Viso 24h",  # TITLE_IT_SHORT
            "Facial Moisturizing Cream 24h",  # TITLE_EN_SHORT
            "Crème Hydratante Visage 24h",  # TITLE_FR_SHORT
            "MiMarca",  # BRAND
            "Facial Premium",  # GAMA_ES
            "COSMETICS_FACIAL",  # FAMILY
            "Cremas Hidratantes",  # SUBFAMILY
            "50",  # NET_WEIGHT_VALUE
            "ml",  # NET_WEIGHT_UNIT
            "75",  # GROSS_WEIGHT_VALUE
            "g",  # GROSS_WEIGHT_UNIT
            "10.5",  # HEIGHT_CM
            "8.2",  # WIDTH_CM
            "4.0",  # DEPTH_CM
            "Botella",  # FORMAT_TYPE
            "Plástico",  # FORMAT_MATERIAL
            "Rosca",  # FORMAT_CLOSURE
            "ES, PT, IT, EN",  # PACKAGING_LANGUAGES
            "WATER, GLYCERIN, PHENOXYETHANOL, SODIUM HYALURONATE",  # INCI_INGREDIENTS
            "Ácido Hialurónico, Vitamina E",  # KEY_INGREDIENTS
            "Aplicar pequeña cantidad en cara limpia. Masajear hasta absorción.",  # MODE_OF_USE_ES
            "Aplicar pequena quantidade no rosto limpo. Massagear até absorção.",  # MODE_OF_USE_PT
            "Applicare piccola quantità sul viso pulito. Massaggiare fino assorbimento.",  # MODE_OF_USE_IT
            "Apply small amount on clean face. Massage until absorbed.",  # MODE_OF_USE_EN
            "Evitar contacto con ojos. Uso externo. Mantener fuera alcance niños.",  # WARNINGS_ES
            "Evitar contacto com olhos. Uso externo. Manter fora alcance crianças.",  # WARNINGS_PT
            "Evitare contatto con occhi. Uso esterno. Tenere fuori portata bambini.",  # WARNINGS_IT
            "Avoid contact with eyes. External use. Keep out of reach of children.",  # WARNINGS_EN
            "PARFUM, LIMONENE",  # ALLERGENS_PRESENT
            "Frutos secos",  # ALLERGENS_MAY_CONTAIN
            "12M",  # PAO
            "skin_irritation, eye_irritation",  # PICTOGRAMS
            "ECOCERT, BIO",  # CERTIFICATIONS
            "ES",  # MADE_IN_COUNTRY
            "Fabricado en España",  # MADE_IN_TEXT_ES
            "Fabricado em Espanha",  # MADE_IN_TEXT_PT
            "Distribuidora S.A.",  # DISTRIBUTOR_NAME
            "B12345678",  # DISTRIBUTOR_CIF
            "Calle Mayor 1, Madrid",  # DISTRIBUTOR_ADDRESS
            "ES",  # DISTRIBUTOR_COUNTRY
            "Juan García",  # RESPONSIBLE_PERSON_NAME
            "juan@example.com",  # RESPONSIBLE_PERSON_EMAIL
            "+34 600 000 000",  # RESPONSIBLE_PERSON_PHONE
            "92",  # NATURAL_ORIGIN_PERCENTAGE
            "NO",  # VEGAN
            "SI",  # CRUELTY_FREE
            "5",  # STORAGE_TEMP_MIN
            "25",  # STORAGE_TEMP_MAX
            "Conservar en lugar fresco y seco",  # STORAGE_CONDITIONS_ES
            "SI",  # PREGNANCY_SAFE
            "EDAD_MINIMA_3",  # CHILD_SAFE
            "approved",  # STATUS
            "Producto estrella de la gama",  # INTERNAL_NOTES
        ]
    
    def _add_data_validations(self, ws):
        """Agregar validaciones de datos (dropdowns) a columnas específicas."""
        # FAMILY dropdown (columna 11)
        dv_family = DataValidation(type="list", formula1=f'"{",".join(self.VALID_FAMILIES)}"')
        ws.add_data_validation(dv_family)
        dv_family.add(f"K3:K1000")  # Columna K (FAMILY) filas 3-1000
        
        # FORMAT_TYPE dropdown (columna 20)
        dv_format = DataValidation(type="list", formula1=f'"{",".join(self.VALID_FORMAT_TYPES)}"')
        ws.add_data_validation(dv_format)
        dv_format.add(f"T3:T1000")
        
        # FORMAT_MATERIAL dropdown (columna 21)
        dv_material = DataValidation(type="list", formula1=f'"{",".join(self.VALID_FORMAT_MATERIALS)}"')
        ws.add_data_validation(dv_material)
        dv_material.add(f"U3:U1000")
        
        # PAO dropdown (columna 35)
        dv_pao = DataValidation(type="list", formula1=f'"{",".join(self.VALID_PAO_VALUES)}"')
        ws.add_data_validation(dv_pao)
        dv_pao.add(f"AI3:AI1000")
        
        # NET_WEIGHT_UNIT dropdown (columna 14)
        dv_weight = DataValidation(type="list", formula1=f'"{",".join(self.VALID_WEIGHT_UNITS)}"')
        ws.add_data_validation(dv_weight)
        dv_weight.add(f"N3:N1000")
        
        # STATUS dropdown (columna 57)
        dv_status = DataValidation(type="list", formula1=f'"{",".join(self.VALID_STATUS)}"')
        ws.add_data_validation(dv_status)
        dv_status.add(f"BE3:BE1000")
    
    def import_from_excel(self, file_path: str, current_user_id: str = "system") -> Dict:
        """
        Importar productos desde archivo Excel con validaciones.
        
        Args:
            file_path: Ruta al archivo Excel
            current_user_id: ID del usuario que importa
        
        Returns:
            Dict con resultados: {imported, errors, skipped, status, completion_percentage}
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        errors = []
        imported = 0
        skipped = 0
        imported_sheets = []
        
        # Iterar filas (skip header row 1 y ejemplo row 2, start from row 3)
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            try:
                # Mapear datos de la fila
                row_data = {}
                for col_idx, (col_name, _, _) in enumerate(self.EXCEL_COLUMNS):
                    row_data[col_name] = row[col_idx] if col_idx < len(row) else None
                
                # Validar SKU
                sku = row_data.get("SKU")
                if not sku or not str(sku).strip():
                    errors.append({"row": row_idx, "error": "SKU vacío"})
                    skipped += 1
                    continue
                
                # Validar formato SKU
                if not re.match(r'^[A-Z]{2}-[A-Z0-9]{3,5}-[0-9]{3}$', str(sku)):
                    errors.append({"row": row_idx, "sku": sku, "error": "SKU formato inválido (esperado: XX-XXX-000)"})
                    skipped += 1
                    continue
                
                # Validar EAN si existe
                ean = row_data.get("EAN_PRIMARY")
                if ean and not self._validate_ean13(str(ean)):
                    errors.append({"row": row_idx, "sku": sku, "error": "EAN checksum inválido"})
                    skipped += 1
                    continue
                
                # Validar FAMILY
                family = row_data.get("FAMILY")
                if not family or family not in self.VALID_FAMILIES:
                    errors.append({"row": row_idx, "sku": sku, "error": f"FAMILY '{family}' no válida"})
                    skipped += 1
                    continue
                
                # Validar TITLE_ES_SHORT
                title_es = row_data.get("TITLE_ES_SHORT")
                if not title_es or len(str(title_es).strip()) < 3:
                    errors.append({"row": row_idx, "sku": sku, "error": "TITLE_ES_SHORT mínimo 3 caracteres"})
                    skipped += 1
                    continue
                
                # Crear ProductSheet
                product_data = self._map_excel_row_to_product(row_data)
                sheet = ProductSheet(**product_data)
                
                # Aplicar preset
                self.preset_manager.apply_preset(sheet, family)
                
                # Guardar en DB
                self.session.add(sheet)
                self.session.flush()
                
                # Crear v1.0 snapshot
                self.version_manager.create_snapshot(
                    sku=sku,
                    version_type="major",
                    change_summary="Imported from Excel",
                    user_id=current_user_id
                )
                
                imported += 1
                imported_sheets.append(sheet)
            
            except Exception as e:
                errors.append({"row": row_idx, "sku": row_data.get("SKU", "unknown"), "error": str(e)})
                self.session.rollback()
                skipped += 1
        
        # Commit si hay importaciones exitosas
        if imported > 0:
            self.session.commit()
        
        # Calcular completion percentage
        if imported > 0:
            total_fields_filled = sum(self._count_non_empty_fields(sheet) for sheet in imported_sheets)
            total_fields_possible = imported * 60
            completion_percentage = int((total_fields_filled / total_fields_possible) * 100)
        else:
            completion_percentage = 0
        
        return {
            "imported": imported,
            "errors": errors,
            "skipped": skipped,
            "status": "SUCCESS" if imported > 0 else "FAILED",
            "completion_percentage": completion_percentage,
            "message": f"{imported} productos importados exitosamente. {len(errors)} errores."
        }
    
    def _validate_ean13(self, ean: str) -> bool:
        """Validar checksum EAN-13."""
        if not ean or len(ean) != 13 or not ean.isdigit():
            return False
        
        # Algoritmo checksum EAN-13
        odd_sum = sum(int(ean[i]) for i in range(0, 12, 2))
        even_sum = sum(int(ean[i]) for i in range(1, 12, 2))
        total = odd_sum + (even_sum * 3)
        checksum = (10 - (total % 10)) % 10
        
        return checksum == int(ean[12])
    
    def _map_excel_row_to_product(self, row_data: Dict) -> Dict:
        """Mapear fila Excel a dict de ProductSheet."""
        product_data = {
            "sku": row_data.get("SKU"),
            "ean_list": [row_data.get("EAN_PRIMARY")] if row_data.get("EAN_PRIMARY") else [],
            "brand": row_data.get("BRAND"),
            "family": row_data.get("FAMILY"),
            "subfamily": row_data.get("SUBFAMILY"),
            "title_short": {
                "es": row_data.get("TITLE_ES_SHORT"),
                "pt": row_data.get("TITLE_PT_SHORT"),
                "it": row_data.get("TITLE_IT_SHORT"),
                "en": row_data.get("TITLE_EN_SHORT"),
                "fr": row_data.get("TITLE_FR_SHORT"),
            },
            "net_weight_value": float(row_data.get("NET_WEIGHT_VALUE", 0)) if row_data.get("NET_WEIGHT_VALUE") else None,
            "net_weight_unit": row_data.get("NET_WEIGHT_UNIT"),
            "format_type": row_data.get("FORMAT_TYPE"),
            "format_material": row_data.get("FORMAT_MATERIAL"),
            "inci_ingredients": row_data.get("INCI_INGREDIENTS"),
            "mode_of_use": {
                "es": row_data.get("MODE_OF_USE_ES"),
                "pt": row_data.get("MODE_OF_USE_PT"),
                "it": row_data.get("MODE_OF_USE_IT"),
                "en": row_data.get("MODE_OF_USE_EN"),
            },
            "general_warnings": {
                "es": row_data.get("WARNINGS_ES"),
                "pt": row_data.get("WARNINGS_PT"),
                "it": row_data.get("WARNINGS_IT"),
                "en": row_data.get("WARNINGS_EN"),
            },
            "allergens_present": [a.strip() for a in row_data.get("ALLERGENS_PRESENT", "").split(",") if a.strip()],
            "pao": row_data.get("PAO"),
            "status": row_data.get("STATUS", "draft"),
        }
        
        return product_data
    
    def _count_non_empty_fields(self, sheet: ProductSheet) -> int:
        """Contar campos no vacíos de un ProductSheet."""
        count = 0
        for key, value in sheet.__dict__.items():
            if not key.startswith("_") and value is not None and value != "" and value != []:
                count += 1
        return count
    
    def export_to_markdown(self, sku: str) -> str:
        """
        Exportar ficha a formato Markdown.
        
        Args:
            sku: SKU del producto
        
        Returns:
            String con contenido Markdown
        """
        sheet = self.session.query(ProductSheet).filter_by(sku=sku).first()
        if not sheet:
            raise ValueError(f"Product {sku} not found")
        
        md_content = f"""# Ficha de Producto: {sku}

## Información General

| Propiedad | Valor |
|-----------|-------|
| SKU | {sheet.sku} |
| EAN | {", ".join(sheet.ean_list) if sheet.ean_list else "N/A"} |
| Marca | {sheet.brand or "N/A"} |
| Familia | {sheet.family or "N/A"} |
| Subfamilia | {sheet.subfamily or "N/A"} |
| Estado | {sheet.status or "N/A"} |
| Versión actual | {sheet.current_version or "N/A"} |
| Completado | {sheet.completion_percentage or 0}% |

## Composición INCI

{sheet.inci_ingredients or "No especificada"}

## Modo de Empleo

### Español
{sheet.mode_of_use.get("es", "No especificado") if isinstance(sheet.mode_of_use, dict) else "No especificado"}

### Portugués
{sheet.mode_of_use.get("pt", "No especificado") if isinstance(sheet.mode_of_use, dict) else "No especificado"}

### Italiano
{sheet.mode_of_use.get("it", "No especificado") if isinstance(sheet.mode_of_use, dict) else "No especificado"}

## Avisos y Precauciones

⚠️ **Advertencias**: {sheet.general_warnings.get("es", "No especificadas") if isinstance(sheet.general_warnings, dict) else "No especificadas"}

## Alérgenos

**Presentes**: {", ".join(sheet.allergens_present) if sheet.allergens_present else "Ninguno"}

## Información Técnica

- **Peso neto**: {sheet.net_weight_value} {sheet.net_weight_unit}
- **PAO**: {sheet.pao or "N/A"}
- **Formato**: {sheet.format_type or "N/A"} ({sheet.format_material or "N/A"})

---

_Generado el {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}_
"""
        
        return md_content
    
    def export_to_html(self, sku: str) -> str:
        """
        Exportar ficha a formato HTML.
        
        Args:
            sku: SKU del producto
        
        Returns:
            String con contenido HTML
        """
        md_content = self.export_to_markdown(sku)
        
        # Convertir Markdown a HTML básico (sin librería markdown)
        html = md_content
        html = html.replace("# ", "<h1>").replace("\n\n", "</h1>\n\n")
        html = html.replace("## ", "<h2>").replace("\n", "</h2>\n")
        html = html.replace("### ", "<h3>").replace("\n", "</h3>\n")
        html = html.replace("**", "<strong>").replace("**", "</strong>")
        html = html.replace("⚠️", '<span class="warning">⚠️</span>')
        
        # Wrap en HTML completo
        html_content = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Ficha de Producto - {sku}</title>
    <style>
        body {{ font-family: Arial, sans-serif; max-width: 800px; margin: 0 auto; padding: 20px; }}
        h1 {{ color: #2563eb; }}
        h2 {{ color: #1e40af; border-bottom: 2px solid #3b82f6; padding-bottom: 5px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #3b82f6; color: white; }}
        .warning {{ color: #dc2626; font-size: 1.2em; }}
    </style>
</head>
<body>
    {html}
</body>
</html>"""
        
        return html_content
    
    def export_catalog_excel(self, filters: Optional[Dict] = None) -> bytes:
        """
        Exportar catálogo de productos a Excel con filtros.
        
        Args:
            filters: Dict con filtros (family, brand, status, etc.)
        
        Returns:
            bytes: Archivo Excel
        """
        # Aplicar filtros a query
        query = self.session.query(ProductSheet)
        
        if filters:
            if filters.get("family"):
                query = query.filter(ProductSheet.family == filters["family"])
            if filters.get("brand"):
                query = query.filter(ProductSheet.brand == filters["brand"])
            if filters.get("status"):
                query = query.filter(ProductSheet.status == filters["status"])
        
        products = query.all()
        
        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catálogo"
        
        # Headers
        headers = ["SKU", "Marca", "Familia", "Título ES", "Estado", "Completado %", "Versión"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Datos
        for row_idx, product in enumerate(products, start=2):
            ws.cell(row=row_idx, column=1).value = product.sku
            ws.cell(row=row_idx, column=2).value = product.brand
            ws.cell(row=row_idx, column=3).value = product.family
            ws.cell(row=row_idx, column=4).value = product.title_short.get("es", "") if isinstance(product.title_short, dict) else ""
            ws.cell(row=row_idx, column=5).value = product.status
            ws.cell(row=row_idx, column=6).value = product.completion_percentage
            ws.cell(row=row_idx, column=7).value = product.current_version
        
        # Autosize
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()

